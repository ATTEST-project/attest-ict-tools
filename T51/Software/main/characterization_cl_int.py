import shutil
from click import command

import xlrd
from openpyxl import load_workbook

import json

import argparse

import re

from chart.lib.main_cluster_maker import *

import config

class ArgParser():
	def __init__(
		self,
    ):
		self.initParser()

	def initParser(self):
		self.parser = argparse.ArgumentParser()
    
	def getParser(self):
		return self.parser

	def addArgument(self, command, type, default):
		self.parser.add_argument(
			"--" + str(command),  # name on the CLI - drop the `--` for positional/required parameters
			type=type, # str/int/float
			default=default,  # default if nothing is provided
		)

	def addArgumentList(self, command, nargs, type, default):
		self.parser.add_argument(
			"--" + str(command),  # name on the CLI - drop the `--` for positional/required parameters
			nargs=nargs,  # 0 or more values expected => creates a list
			type=type, # str/int/float
			default=default,  # default if nothing is provided
		)

class T511Tool():
	def __init__(
            self,
            argsParsed
    ):
		self.configs = argsParsed

	def getConfigs(self):
		return self.configs

	def generate_results(self):
		print("Generating Result")
		input_data(self.configs)

	def readCSVFile(self, config):
		# print("Config from parse method: " + str(config))
		filePath = config["path"]
		components = []
		componentIndex = ""

		print("\nReading csv file...")

		# csv_reader = pd.read_csv(config["path"])

		csv_header_1 = pd.read_csv(filePath, sep=";",index_col=0, nrows=0)
		csv_header_2 = pd.read_csv(filePath, index_col=0, nrows=0)
		csv_reader = pd.read_csv(filePath, sep=";") if csv_header_1.columns.__len__() >= csv_header_2.columns.__len__() else pd.read_csv(filePath)

		headers = csv_reader.columns

		for header in headers:
			if header == config["selectedVariables"]:
				componentIndex = header

		cnt = len(csv_reader)
		print(cnt)
		for i in range(0, cnt):
			if not csv_reader[componentIndex][i] in components:
				# print(csv_reader[componentIndex][i])
				components.append(csv_reader[componentIndex][i])

		print("\nReading csv file completed")

		return components

	def readXLSFile(self, config):
		filePath = config["path"]
		components = []
		componentIndex = 0

		print("\nReading xls file...")

		wb = xlrd.open_workbook(filePath)
		sheet = wb.sheet_by_index(0)
		rowCount = sheet.nrows
		colCount = sheet.ncols
		for row in range(rowCount):
			if row == 0:
				index = 0
				headerTexts = []
				for col in range(colCount):
					header = str(sheet.cell_value(row, col))
					if header != "":
						headerTexts.append(header)

				for col in range(colCount):
					header = str(sheet.cell_value(row, col))
					if header == config["component1_field"]:
						componentIndex = index
					index = index + 1
			else:
				if not sheet.cell_value(row, componentIndex) in components:
					components.append(sheet.cell_value(row, componentIndex))

		print("\nReading xls file completed")
		return components

	def readXLSXFile(self, config):
		filePath = config["path"]
		components = []
		componentIndex = 0

		print("\nReading xlsx file...")
		wb = load_workbook(filename=filePath)
		sheet = wb.active
		rowCount = sheet.max_row
		colCount = sheet.max_column
		for row in range(1, rowCount + 1):
			if row == 1:
				index = 1
				headerTexts = []
				for col in range(1, colCount + 1):
					header = str(sheet.cell(row=row, column=col).value)
					if header != "":
						headerTexts.append(header)

				for col in range(1, colCount + 1):
					header = str(sheet.cell(row=row, column=col).value)
					if header == config["component1_field"]:
						componentIndex = index
					index = index + 1
			else:
				if not sheet.cell(row=row, column=componentIndex).value in components:
					components.append(str(sheet.cell(row=row, column=componentIndex).value))

		print("\nReading xlsx file completed")
		return components

	def readCSVAuxFile(self, config):
		# print("Config from parse method: " + str(config))
		filePath = config["path2"]
		components = []
		componentIndex = ""

		print("\nReading csv file...")

		# csv_reader = pd.read_csv(config["path"])

		csv_header_1 = pd.read_csv(filePath, sep=";",index_col=0, nrows=0)
		csv_header_2 = pd.read_csv(filePath, index_col=0, nrows=0)
		csv_reader = pd.read_csv(filePath, sep=";") if csv_header_1.columns.__len__() >= csv_header_2.columns.__len__() else pd.read_csv(filePath)

		headers = csv_reader.columns

		for header in headers:
			if header == config["variables2"]:
				componentIndex = header

		cnt = len(csv_reader)
		print(cnt)
		for i in range(0, cnt):
			if not csv_reader[componentIndex][i] in components:
				# print(csv_reader[componentIndex][i])
				components.append(csv_reader[componentIndex][i])

		print("\nReading csv file completed")

		return components

	def readXLSAuxFile(self, config):
		filePath = config["path"]
		components = []
		componentIndex = 0

		print("\nReading xls file...")

		wb = xlrd.open_workbook(filePath)
		sheet = wb.sheet_by_index(0)
		rowCount = sheet.nrows
		colCount = sheet.ncols
		for row in range(rowCount):
			if row == 0:
				index = 0
				headerTexts = []
				for col in range(colCount):
					header = str(sheet.cell_value(row, col))
					if header != "":
						headerTexts.append(header)

				for col in range(colCount):
					header = str(sheet.cell_value(row, col))
					if header == config["component2_field"]:
						componentIndex = index
					index = index + 1
			else:
				if not sheet.cell_value(row, componentIndex) in components:
					components.append(sheet.cell_value(row, componentIndex))

		print("\nReading xls file completed")
		return components

	def readXLSXAuxFile(self, config):
		filePath = config["path"]
		components = []
		componentIndex = 0

		print("\nReading xlsx file...")
		wb = load_workbook(filename=filePath)
		sheet = wb.active
		rowCount = sheet.max_row
		colCount = sheet.max_column
		for row in range(1, rowCount + 1):
			if row == 1:
				index = 1
				headerTexts = []
				for col in range(1, colCount + 1):
					header = str(sheet.cell(row=row, column=col).value)
					if header != "":
						headerTexts.append(header)

				for col in range(1, colCount + 1):
					header = str(sheet.cell(row=row, column=col).value)
					if header == config["component2_field"]:
						componentIndex = index
					index = index + 1
			else:
				if not sheet.cell(row=row, column=componentIndex).value in components:
					components.append(str(sheet.cell(row=row, column=componentIndex).value))

		print("\nReading xlsx file completed")
		return components

	def setComponents(self):
		# print("\nConfigs from loop method: " + str(self.configs))
		for i in range(self.configs.__len__()):
			# print("\n" + str(self.configs[i]))
			filePath = self.configs[i]["path"]
			file_extension = os.path.splitext(filePath)[1]
			# print("\nFilePath: " + str(filePath) + " with file extension: " + str(file_extension))
			if file_extension == ".csv":
				self.configs[i]["components"] = self.readCSVFile(self.configs[i])
			elif file_extension == ".xls":
				self.configs[i]["components"] = self.readXLSFile(self.configs[i])
			else:
				self.configs[i]["components"] = self.readXLSXFile(self.configs[i])

			if not self.configs[i]["path2"] == "":
				auxFilePath = self.configs[i]["path2"]
				auxFileExtension = os.path.splitext(auxFilePath)[1]
				if auxFileExtension == ".csv":
					self.configs[i]["components2"] = self.readCSVAuxFile(self.configs[i])
				elif auxFileExtension == ".xls":
					self.configs[i]["components2"] = self.readXLSAuxFile(self.configs[i])
				else:
					self.configs[i]["components2"] = self.readXLSXAuxFile(self.configs[i])

def startArgsParser():
	args = ArgParser()
	args.addArgument("json", str, "[]")
	args.addArgument("json-file", str, "")
	parser = args.getParser()
	argsParsed = parser.parse_args()
	return argsParsed

def main():
	argsParsed = startArgsParser()
	jsonFile = open(argsParsed.json_file)
	jsonParsed = json.load(jsonFile)
	# print("\nJSON Parsed: " + str(jsonParsed))
	config.output_dir_from_config = jsonParsed['output_dir']
	app = T511Tool(jsonParsed['configs'])
	app.setComponents()
	# print("updated configs: " + str(app.getConfigs()))
	app.generate_results()

if __name__ == '__main__':
    main()